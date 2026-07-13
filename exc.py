from openpyxl import load_workbook
from datetime import datetime

RUTA_EXCEL = r"https://docs.google.com/spreadsheets/d/1BbHMpnBPXbVzN3QMLObL7CDn2QDyUYJI/edit?usp=sharing&ouid=114112168677691215677&rtpof=true&sd=true"

HOJA = "ASISTENCIA"

GRUPOS = {
    "A": {
        "fila_inicio": 8,
        "fila_fin": 22,
        "fila_fecha": 7
    },
    "B": {
        "fila_inicio": 46,
        "fila_fin": 60,
        "fila_fecha": 45
    },
    "C": {
        "fila_inicio": 86,
        "fila_fin": 101,
        "fila_fecha": 85
    }
}


def GuardarAsistencia(grupo, estudiantes, asistencias):

    libro = load_workbook(RUTA_EXCEL)
    hoja = libro[HOJA]

    fila_inicio = GRUPOS[grupo]["fila_inicio"]
    fila_fecha = GRUPOS[grupo]["fila_fecha"]

    # Buscar la primera columna vacía entre C y M
    columna = None

    for c in range(3, 14):      # C=3 hasta M=13

        vacia = True

        for fila in range(
            GRUPOS[grupo]["fila_inicio"],
            GRUPOS[grupo]["fila_fin"] + 1
        ):

            if hoja.cell(row=fila, column=c).value is not None:

                vacia = False
                break

        if vacia:

            columna = c
            break

    if columna is None:
        raise Exception("Ya no quedan columnas disponibles (C hasta M).")

    # Escribir la fecha en el encabezado
    hoja.cell(
        row=fila_fecha,
        column=columna
    ).value = datetime.now().strftime("%d/%m/%Y")

    # Guardar las asistencias
    fila = fila_inicio

    for estudiante in estudiantes:

        estado = asistencias[estudiante["numero"]]

        if estado == "Presente":
            hoja.cell(row=fila, column=columna).value = "."

        elif estado == "Tardanza":
            hoja.cell(row=fila, column=columna).value = "T"

        else:
            hoja.cell(row=fila, column=columna).value = "---"

        fila += 1

    libro.save(RUTA_EXCEL)
